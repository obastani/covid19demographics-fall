# See CO.py for pydrive setup

import csv
from datetime import datetime
from openpyxl import load_workbook
import os
import pydrive
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import re
import pandas as pd

def run_RI(args):
    # Parameters
    raw_name = '../RI/raw'
    data_name = '../RI/data/data.csv'
    now = str(datetime.now())

    # Reading from downloaded file - HERE change this below to access the sheet
    # try:
    #     demo_df = pd.read_excel(raw_name + '/2020-04-21 182700.723233.xlsx', 
    #                         sheet_name='Demographics')
    # except:
    #     print('Error Reading Sheet Demographics')
    #     raise

    # Download current spreadsheet from gdrive to raw folder
    gauth = GoogleAuth()
    gauth.LocalWebserverAuth()
    drive = GoogleDrive(gauth)
    fileid = '1c2QrNMz8pIbYEKzMJL7Uh2dtThOJa2j1sSMwiDo5Gz4'
    fp = drive.CreateFile({"id": fileid})
    fname = "%s/%s.xlsx" % (raw_name, now)
    fp.GetContentFile(fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    demo_df = pd.read_excel(fname, sheet_name='Demographics')

    # print(demo_df)
    # exit() 

    # Expected column names
    sex_cols = ['Female', 'Male', 'Other', 'Pending further information']
    age_cols = ['0-4', '5-9', '10-14', '15-18', '19-24', '25-29', '30-39', '40-49', '50-59', '60-69', '70-79', '80-89', '90+', "Pending further information"]
    race_cols = ["Hispanic or Latino†",
                 "American Indian or Alaska Native*",
                 "Asian*",
                 "Black or African American*",
                 "Native Hawaiian or Other Pacific Islander*",
                 "White*",
                 "Other race*",
                 "Multiple race*",
                 "Declined",
                 "Unknown or pending further information"]
    categories = ['Cases', 'Hospitalizations', 'Deaths']

    # Check categories are valid
    for cat in categories:
        if any(cat in col for col in demo_df.columns):
            continue
        else:
            raise Exception('Check Column Names:' + str(categories))
    
    cases_str = [i for i in demo_df.columns if categories[0] in i][0]
    hosp_str = [i for i in demo_df.columns if categories[1] in i][0]
    fatalities_str = [i for i in demo_df.columns if categories[2] in i][0]

    # Variables to collect data
    sex_data_cases = []
    age_data_cases = []
    race_data_cases = []
    sex_data_hosp = []
    age_data_hosp = []
    race_data_hosp = []
    sex_data_deaths = []
    age_data_deaths = []
    race_data_deaths = []

    # Loop through entire sheet and extract information to above variables
    for i in range(len(demo_df)):
        if demo_df.iloc[i, 0] == 'Sex':
            # Checks last element in sex_cols matches the last sex cat in raw
            if demo_df.iloc[i + len(sex_cols), 0] != sex_cols[-1]:
                raise Exception("Sex Categories do not match")
            # Appends values for each category for cases, hosps and deaths
            for k in range(len(sex_cols)):
                sex_data_cases.append(demo_df.loc[i + k + 1, cases_str])
                sex_data_hosp.append(demo_df.loc[i + k + 1, hosp_str])
                sex_data_deaths.append(demo_df.loc[i + k + 1, fatalities_str])
        elif demo_df.iloc[i, 0] == "Age Group":
            age_val = age_cols[-1]
            if isinstance(age_val, datetime):
                age_val = str(age_val.month) + "-" + str(age_val.day)
            # Checks last element in age_cols matches the last age cat in raw
            if demo_df.iloc[i + len(age_cols), 0] != age_val:
                print(demo_df.iloc[i + len(age_cols), 0], age_cols[-1])
                raise Exception("Age Categories do not match: " + demo_df.iloc[i + len(age_cols), 0])
            # Appends values for each category for cases, hosps and deaths
            for k in range(len(age_cols)):
                age_data_cases.append(demo_df.loc[i + k + 1, cases_str])
                age_data_hosp.append(demo_df.loc[i + k + 1, hosp_str])
                age_data_deaths.append(demo_df.loc[i + k + 1, fatalities_str])
        elif demo_df.iloc[i, 0] == "Race and Ethnicity":
            # Checks last element in age_cols matches the last age cat in raw
            if demo_df.iloc[i + len(race_cols), 0] != race_cols[-1]:
                print(demo_df)
                print(demo_df.iloc[i + len(race_cols), 0], race_cols[-1])
                raise Exception("Race Categories do not match: " )
            # Appends values for each category for cases, hosps and deaths
            for k in range(len(race_cols)):
                race_data_cases.append(demo_df.loc[i + k + 1, cases_str])
                race_data_hosp.append(demo_df.loc[i + k + 1, hosp_str])
                race_data_deaths.append(demo_df.loc[i + k + 1, fatalities_str])
    
    # Create Dictionaries for easy access to data
    sex_dict = {}
    age_dict = {}
    race_dict = {}

    # Add values to dictionary - structure is a dict[sex] to dict[cases/hosp/deaths]
    for idx in range(len(sex_cols)):
        sex_dict[sex_cols[idx]] = {
            'Cases': sex_data_cases[idx],
            'Hospitalizations': sex_data_hosp[idx],
            'Deaths': sex_data_deaths[idx]
        }

    # Add values to dictionary - structure is a dict[age] to dict[cases/hosp/deaths]
    for idx in range(len(age_cols)):
        age_dict[age_cols[idx]] = {
            'Cases': age_data_cases[idx],
            'Hospitalizations': age_data_hosp[idx],
            'Deaths': age_data_deaths[idx]
        }

    # Add values to dictionary - structure is a dict[age] to dict[cases/hosp/deaths]
    for idx in range(len(race_cols)):
        race_dict[race_cols[idx]] = {
            'Cases': race_data_cases[idx],
            'Hospitalizations': race_data_hosp[idx],
            'Deaths': race_data_deaths[idx]
        }
    
    # Create out dictionary
    out = {}
    for age in age_dict:
        for cat in age_dict[age]:
            out["# " + cat + " Age [" + age + "]"] = age_dict[age][cat]

    for sex in sex_dict:
        for cat in sex_dict[sex]:
            out[cat + "_Sex: " + sex] = sex_dict[sex][cat]
    
    for race in race_dict:
        for cat in race_dict[race]:
            out[cat + "_race/ethn: " + race] = race_dict[race][cat]

    # Write to file
    fields = sorted([x for x in out])
    exists = os.path.exists(data_name)
    with open(data_name, "a") as fp:
        writer = csv.writer(fp)
        if not exists:
            writer.writerow(fields)
        writer.writerow([out[x] for x in fields])
    

# def run_RI(args):
    
#     # Parameters
#     raw_name = '../RI/raw'
#     data_name = '../RI/data/data.csv'
#     now = str(datetime.now())
#     # Download current spreadsheet from gdrive to raw folder
#     gauth = GoogleAuth()
#     gauth.LocalWebserverAuth()
#     drive = GoogleDrive(gauth)
#     fileid = '1n-zMS9Al94CPj_Tc3K7Adin-tN9x1RSjjx2UzJ4SV7Q'
#     fp = drive.CreateFile({"id": fileid})
#     fname = "%s/%s.xlsx" % (raw_name, now)
#     fp.GetContentFile(fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#     # Pull data from workbook to csv
#     out = {}
#     wb = load_workbook(filename=fname)
#     # HERE - new worksheet name is "Demographics"
#     if "Age" not in wb or "Sex" not in wb:
#         raise Exception("Missing sheet")

#     # age
#     cols = [list(x) for x in wb["Age"].iter_cols(values_only=True)]
#     for idx in range(len(cols[0])):
#         if type(cols[0][idx]).__name__ == "datetime":
#             cols[0][idx] = "10-19"
#     ages = ['0-9', '10-19', '20-29', '30-39', '40-49', '50-59', '60-69', '70-79', '80-89', '90-99', '100+']
#     if len(cols) != 3 or tuple([x.strip() for x in cols[0] if x is not None]) != tuple(["Age Group"] + ages):
#         raise Exception("Unexpected age ranges")
#     for idx in range(len(ages)):
#         field = "Age_" + ages[idx].replace("-", "_").replace("+", "_plus")
#         if isinstance(cols[1][idx+1], str) and "<" in cols[1][idx+1]:
#             out[field] = None
#         else:
#             out[field] = int(cols[1][idx+1])
#     reupdate = re.compile("updated\s+(\d+/\d+/\d+)")
#     ageupdate = reupdate.search(cols[2][0])
#     if not ageupdate:
#         raise Exception("Unexpected age update string")
#     spl = ageupdate.group(1).split("/")
#     out["UpdateAge"] = spl[2] + "-" + spl[0].zfill(2) + "-" + spl[1].zfill(2)

#     # gender
#     cols = [x for x in wb["Sex"].iter_cols(values_only=True)]
#     if len(cols) != 3 or cols[0] != ("Sex", "Female", "Male", "Other", None):
#         raise Exception("Unexpected gender cats")
#     out["CasesFemale"] = int(cols[1][1])
#     out["CasesMale"] = int(cols[1][2])
#     out["CasesOther"] = int(cols[1][3])
#     genderupdate = reupdate.search(cols[2][0])
#     if not genderupdate:
#         raise Exception("Unexpected gender update string")
#     spl = genderupdate.group(1).split("/")
#     out["UpdateGender"] = spl[2] + "-" + spl[0].zfill(2) + "-" + spl[1].zfill(2)

#     # output
#     out["Scrape_Time"] = now
#     fields = sorted([x for x in out])
#     exists = os.path.exists(data_name)
#     with open(data_name, "a") as fp:
#         writer = csv.writer(fp)
#         if not exists:
#             writer.writerow(fields)
#         writer.writerow([out[x] for x in fields])

if __name__ == '__main__':
    run_RI({})
